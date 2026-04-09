from pathlib import Path

from openpyxl import load_workbook

from src.exporter import export_month_excel
from src.ledger_store import LedgerStore
from src.transaction_parser import ParsedTransaction


def test_export_month_excel(tmp_path: Path):
    db_path = tmp_path / "ledger.db"
    exports_dir = tmp_path / "exports"
    store = LedgerStore(db_path)

    store.add_transaction(
        ParsedTransaction(
            date="2026-04-08",
            transaction_amount=20.0,
            category="chocolate",
            extra_tags="",
        )
    )
    store.add_transaction(
        ParsedTransaction(
            date="2026-05-01",
            transaction_amount=100.0,
            category="food",
            extra_tags="swiggy",
        )
    )

    output = export_month_excel(store, "2026-04", exports_dir)
    assert output.suffix == ".xlsx"

    workbook = load_workbook(output)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("date", "transaction_amount", "category", "extra_tags")
    assert ("2026-04-08", 20, "chocolate", None) in rows
    assert ("2026-05-01", 100, "food", "swiggy") not in rows
